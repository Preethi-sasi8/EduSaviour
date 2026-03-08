# ═══════════════════════════════════════════════════════════════
# page_cohort.py  —  Cohort Comparison (dedicated page)
# ═══════════════════════════════════════════════════════════════

import streamlit as st
import openpyxl
from openpyxl.utils import get_column_letter

EXCEL_PATH = "cohort_data.xlsx"

# ── Field config (key, label, higher_is_better, max_val) ───────
FIELDS = [
    ("Attendance_Percentage",        "Attendance %",            True,  100),
    ("Previous_Year_Marks",          "Previous Year Marks",     True,  100),
    ("Failed_Subjects_Count",        "Failed Subjects",         False, 5),
    ("Distance_to_School_km",        "Distance to School (km)", False, 30),
    ("Siblings_Dropout_Count",       "Siblings Dropout",        False, 3),
    ("Sibling_Dropout_Impact_Score", "Sibling Impact Score",    False, 10),
    ("No_of_Siblings",               "No. of Siblings",         False, 6),
    ("Dropout_Risk_Score",           "Dropout Risk Score",      False, 100),
    ("Age",                          "Age",                     False, 18),
]


# ── Load class profiles from Excel ────────────────────────────
@st.cache_data
def load_profiles():
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb["Class_Profiles"]
        headers = [cell.value for cell in ws[1]]
        data = {}   # {standard: {"Average": {...}, "Topper": {...}}}
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(headers, row))
            std  = int(d["Standard"])
            ptype = d["Profile_Type"]   # "Average" or "Topper"
            if std not in data:
                data[std] = {}
            data[std][ptype] = {k: float(v) if isinstance(v, (int, float)) else v
                                for k, v in d.items()}
        return data
    except Exception as e:
        return None


# ── Load individual students for scatter context ───────────────
@st.cache_data
def load_students(standard):
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb["Individual_Students"]
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(headers, row))
            if int(d["Standard"]) == int(standard):
                rows.append(d)
        return rows
    except:
        return []


# ── Colour logic ───────────────────────────────────────────────
def bar_color(s_val, a_val, t_val, higher_better):
    if higher_better:
        if s_val >= t_val * 0.92:   return "#22c55e", "🟢", "Near Topper"
        elif s_val >= a_val:        return "#3b82f6", "🔵", "Above Average"
        elif s_val >= a_val * 0.8:  return "#f59e0b", "🟡", "Below Average"
        else:                       return "#ef4444", "🔴", "Needs Attention"
    else:
        if s_val <= a_val * 0.6:    return "#22c55e", "🟢", "Excellent"
        elif s_val <= a_val:        return "#3b82f6", "🔵", "At Average"
        elif s_val <= a_val * 1.3:  return "#f59e0b", "🟡", "Slightly High"
        else:                       return "#ef4444", "🔴", "Needs Attention"


# ── Single field bar row ───────────────────────────────────────
def field_bar(label, s_val, a_val, t_val, higher_better, max_val):
    color, icon, status = bar_color(s_val, a_val, t_val, higher_better)

    def pct(v):
        return min(round(float(v) / max_val * 100, 1), 100) if max_val else 0

    st.markdown(f"""
    <div style='background:#0f172a;border-radius:12px;padding:14px 18px;margin-bottom:10px;
                border:1px solid #1e293b;'>
      <div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;'>
        <span style='font-size:14px;font-weight:700;color:#e2e8f0;'>{label}</span>
        <span style='background:{color}22;border:1px solid {color};border-radius:20px;
                     padding:2px 10px;font-size:12px;font-weight:700;color:{color};'>
          {icon} {status}
        </span>
      </div>
      <!-- Student bar -->
      <div style='display:flex;align-items:center;gap:8px;margin-bottom:5px;'>
        <span style='font-size:11px;color:#94a3b8;width:80px;'>This Student</span>
        <div style='flex:1;background:#1e293b;border-radius:6px;height:14px;'>
          <div style='background:{color};width:{pct(s_val)}%;height:14px;border-radius:6px;'></div>
        </div>
        <span style='font-size:12px;font-weight:700;color:{color};width:40px;text-align:right;'>{s_val}</span>
      </div>
      <!-- Average bar -->
      <div style='display:flex;align-items:center;gap:8px;margin-bottom:5px;'>
        <span style='font-size:11px;color:#94a3b8;width:80px;'>Class Avg</span>
        <div style='flex:1;background:#1e293b;border-radius:6px;height:10px;'>
          <div style='background:#64748b;width:{pct(a_val)}%;height:10px;border-radius:6px;'></div>
        </div>
        <span style='font-size:12px;color:#94a3b8;width:40px;text-align:right;'>{a_val}</span>
      </div>
      <!-- Topper bar -->
      <div style='display:flex;align-items:center;gap:8px;'>
        <span style='font-size:11px;color:#94a3b8;width:80px;'>Topper</span>
        <div style='flex:1;background:#1e293b;border-radius:6px;height:10px;'>
          <div style='background:#f59e0b;width:{pct(t_val)}%;height:10px;border-radius:6px;'></div>
        </div>
        <span style='font-size:12px;color:#f59e0b;width:40px;text-align:right;'>{t_val}</span>
      </div>
    </div>
    """, unsafe_allow_html=True)
    return color, status


# ── Class rank card ────────────────────────────────────────────
def rank_card(student, students_list):
    if not students_list:
        return
    try:
        s_marks = float(student.get("Previous_Year_Marks", 0))
        s_att   = float(student.get("Attendance_Percentage", 0))
        # Rank by marks
        sorted_marks = sorted(
            [float(r.get("Previous_Year_Marks", 0)) for r in students_list], reverse=True
        )
        rank = next((i+1 for i, m in enumerate(sorted_marks) if m <= s_marks), len(sorted_marks))
        total = len(students_list)
        pct   = round((1 - rank/total) * 100)

        # Count better/same/worse
        above_avg_count = sum(1 for r in students_list
                              if float(r.get("Previous_Year_Marks", 0)) >= s_marks)

        st.markdown(f"""
        <div style='display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;margin-bottom:20px;'>
          <div style='background:#1e3a5f;border-radius:12px;padding:16px;text-align:center;'>
            <div style='font-size:32px;font-weight:800;color:#60a5fa;'>#{rank}</div>
            <div style='font-size:12px;color:#94a3b8;margin-top:4px;'>Class Rank (Marks)</div>
            <div style='font-size:11px;color:#64748b;'>out of {total} students</div>
          </div>
          <div style='background:#14532d22;border:1px solid #22c55e44;border-radius:12px;
                      padding:16px;text-align:center;'>
            <div style='font-size:32px;font-weight:800;color:#22c55e;'>Top {100-pct}%</div>
            <div style='font-size:12px;color:#94a3b8;margin-top:4px;'>Percentile Position</div>
            <div style='font-size:11px;color:#64748b;'>based on marks</div>
          </div>
          <div style='background:#1e293b;border-radius:12px;padding:16px;text-align:center;'>
            <div style='font-size:32px;font-weight:800;color:#f59e0b;'>{total}</div>
            <div style='font-size:12px;color:#94a3b8;margin-top:4px;'>Total Students</div>
            <div style='font-size:11px;color:#64748b;'>in this class dataset</div>
          </div>
        </div>
        """, unsafe_allow_html=True)
    except Exception:
        pass


# ══════════════════════════════════════════════════════════════
# MAIN SHOW
# ══════════════════════════════════════════════════════════════
def show():
    student = st.session_state.get("student_data", {})
    result  = st.session_state.get("result", {})

    # ── Header ─────────────────────────────────────────────────
    st.title("📊 Cohort Comparison")

    # Back button top-right
    col_title, col_back = st.columns([5, 1])
    with col_back:
        if st.button("⬅️ Back", use_container_width=True):
            st.session_state.page = "main"
            st.rerun()

    # ── No data guard ───────────────────────────────────────────
    if not student or not result:
        st.markdown("""
        <div style='background:#1e293b;border-radius:14px;padding:40px;text-align:center;
                    border:2px dashed #334155;margin-top:20px;'>
          <div style='font-size:48px;margin-bottom:12px;'>📋</div>
          <div style='font-size:20px;font-weight:700;color:#f1f5f9;'>No Student Data Found</div>
          <div style='font-size:14px;color:#94a3b8;margin-top:8px;'>
            Please go to Risk Prediction, fill in student details and click
            <b style='color:#3b82f6;'>Predict Dropout Risk</b> first.
          </div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🏠 Go to Risk Prediction", type="primary"):
            st.session_state.page = "main"
            st.rerun()
        return

    # ── Load cohort data ────────────────────────────────────────
    profiles = load_profiles()
    standard = int(student.get("Standard", 9))

    if not profiles or standard not in profiles:
        st.error(f"❌ Could not load cohort data for Class {standard}. "
                 f"Make sure **cohort_data.xlsx** is in the project folder.")
        return

    avg_profile    = profiles[standard]["Average"]
    topper_profile = profiles[standard]["Topper"]
    students_list  = load_students(standard)

    # ── Student banner ──────────────────────────────────────────
    name = student.get("name", "Student")
    sid  = student.get("id", "—")
    risk = result.get("Risk_Level", "Unknown")
    risk_colors = {
        "High":     ("#7f1d1d", "#ef4444", "🔴"),
        "Moderate": ("#78350f", "#f59e0b", "🟠"),
        "Low":      ("#14532d", "#22c55e", "🟢"),
    }
    rbg, rborder, ricon = risk_colors.get(risk, ("#1e293b", "#64748b", "⚪"))

    st.markdown(f"""
    <div style='background:{rbg};border-left:6px solid {rborder};border-radius:14px;
                padding:18px 24px;margin-bottom:20px;
                display:flex;justify-content:space-between;align-items:center;'>
      <div>
        <div style='font-size:22px;font-weight:800;color:#fff;'>👤 {name}</div>
        <div style='font-size:13px;color:#ccc;margin-top:4px;'>
          ID: {sid} &nbsp;|&nbsp; Class {standard}
        </div>
      </div>
      <div style='font-size:24px;font-weight:800;color:#fff;'>{ricon} {risk} Risk</div>
    </div>
    """, unsafe_allow_html=True)

    # ── Rank cards ──────────────────────────────────────────────
    rank_card(student, students_list)

    # ── Legend ─────────────────────────────────────────────────
    st.markdown("""
    <div style='display:flex;gap:16px;flex-wrap:wrap;margin-bottom:16px;'>
      <span style='display:flex;align-items:center;gap:6px;font-size:13px;color:#e2e8f0;'>
        <span style='width:28px;height:10px;background:#22c55e;border-radius:3px;display:inline-block;'></span>🟢 Near Topper / Excellent
      </span>
      <span style='display:flex;align-items:center;gap:6px;font-size:13px;color:#e2e8f0;'>
        <span style='width:28px;height:10px;background:#3b82f6;border-radius:3px;display:inline-block;'></span>🔵 Above / At Average
      </span>
      <span style='display:flex;align-items:center;gap:6px;font-size:13px;color:#e2e8f0;'>
        <span style='width:28px;height:10px;background:#f59e0b;border-radius:3px;display:inline-block;'></span>🟡 Slightly Below
      </span>
      <span style='display:flex;align-items:center;gap:6px;font-size:13px;color:#e2e8f0;'>
        <span style='width:28px;height:10px;background:#ef4444;border-radius:3px;display:inline-block;'></span>🔴 Needs Attention
      </span>
    </div>
    <div style='display:flex;gap:16px;flex-wrap:wrap;margin-bottom:20px;'>
      <span style='font-size:12px;color:#94a3b8;'>
        Bar sizes: <b style='color:#94a3b8;'>████</b> Student &nbsp;
        <b style='color:#64748b;'>███</b> Class Average &nbsp;
        <b style='color:#f59e0b;'>███</b> Topper
      </span>
    </div>
    """, unsafe_allow_html=True)

    # ── Field bars ─────────────────────────────────────────────
    st.subheader(f"📈 Field-by-Field Comparison — Class {standard}")
    st.caption(f"Comparing against Class {standard} average and topper from cohort_data.xlsx")

    summary_colors = []
    for key, label, higher_better, max_val in FIELDS:
        try:
            s_val = round(float(student.get(key, 0) or 0), 1)
        except (TypeError, ValueError):
            s_val = 0.0
        a_val = round(avg_profile.get(key, 0), 1)
        t_val = round(topper_profile.get(key, 0), 1)
        color, status = field_bar(label, s_val, a_val, t_val, higher_better, max_val)
        summary_colors.append((label, color, status, s_val, a_val))

    st.divider()

    # ── Summary scorecard ───────────────────────────────────────
    st.subheader("🧾 Performance Summary")

    green  = sum(1 for _, c, _, _, _ in summary_colors if c == "#22c55e")
    blue   = sum(1 for _, c, _, _, _ in summary_colors if c == "#3b82f6")
    yellow = sum(1 for _, c, _, _, _ in summary_colors if c == "#f59e0b")
    red    = sum(1 for _, c, _, _, _ in summary_colors if c == "#ef4444")
    total  = len(summary_colors)

    st.markdown(f"""
    <div style='display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:12px;margin-bottom:20px;'>
      <div style='background:#14532d22;border:2px solid #22c55e44;border-radius:12px;
                  padding:16px;text-align:center;'>
        <div style='font-size:36px;font-weight:800;color:#22c55e;'>{green}</div>
        <div style='font-size:12px;color:#86efac;margin-top:4px;'>🟢 Near Topper / Excellent</div>
      </div>
      <div style='background:#1e3a5f22;border:2px solid #3b82f644;border-radius:12px;
                  padding:16px;text-align:center;'>
        <div style='font-size:36px;font-weight:800;color:#3b82f6;'>{blue}</div>
        <div style='font-size:12px;color:#93c5fd;margin-top:4px;'>🔵 Above / At Average</div>
      </div>
      <div style='background:#78350f22;border:2px solid #f59e0b44;border-radius:12px;
                  padding:16px;text-align:center;'>
        <div style='font-size:36px;font-weight:800;color:#f59e0b;'>{yellow}</div>
        <div style='font-size:12px;color:#fcd34d;margin-top:4px;'>🟡 Slightly Below</div>
      </div>
      <div style='background:#7f1d1d22;border:2px solid #ef4444 44;border-radius:12px;
                  padding:16px;text-align:center;'>
        <div style='font-size:36px;font-weight:800;color:#ef4444;'>{red}</div>
        <div style='font-size:12px;color:#fca5a5;margin-top:4px;'>🔴 Needs Attention</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Key insights ────────────────────────────────────────────
    insights = []
    s_marks = float(student.get("Previous_Year_Marks", 0) or 0)
    s_att   = float(student.get("Attendance_Percentage", 0) or 0)
    a_marks = avg_profile.get("Previous_Year_Marks", 0)
    a_att   = avg_profile.get("Attendance_Percentage", 0)
    t_marks = topper_profile.get("Previous_Year_Marks", 0)

    if s_marks < a_marks:
        insights.append(f"📉 Marks are <b>{round(a_marks - s_marks, 1)} pts below</b> class average — needs academic support")
    else:
        insights.append(f"📈 Marks are <b>{round(s_marks - a_marks, 1)} pts above</b> class average — keep it up!")

    gap = round(t_marks - s_marks, 1)
    if gap > 0:
        insights.append(f"🏆 Needs <b>{gap} more marks</b> to reach topper level")
    else:
        insights.append(f"🌟 Performance is <b>at topper level</b> — excellent!")

    if s_att < a_att:
        insights.append(f"⚠️ Attendance is <b>{round(a_att - s_att, 1)}% below</b> class average — critical to improve")
    else:
        insights.append(f"✅ Attendance is <b>{round(s_att - a_att, 1)}% above</b> class average — great consistency")

    if red >= 3:
        insights.append(f"🚨 <b>{red} fields</b> need immediate attention — consider counselling")
    elif green + blue >= 6:
        insights.append(f"🎯 <b>{green + blue} out of {total} fields</b> are performing well")

    bullets = "".join([f"<li style='margin-bottom:6px;'>{i}</li>" for i in insights])
    st.markdown(f"""
    <div style='background:#1e293b;border-radius:12px;padding:18px 22px;
                border-left:5px solid #3b82f6;'>
      <div style='font-size:15px;font-weight:700;color:#93c5fd;margin-bottom:10px;'>
        💡 Key Insights
      </div>
      <ul style='color:#cbd5e1;font-size:14px;margin:0;padding-left:20px;line-height:1.8;'>
        {bullets}
      </ul>
    </div>
    """, unsafe_allow_html=True)

    st.divider()

    # ── Bottom navigation ───────────────────────────────────────
    c1, c2, c3 = st.columns(3)
    with c1:
        if st.button("⬅️ Back to Prediction", use_container_width=True):
            st.session_state.page = "main"; st.rerun()
    with c2:
        if st.button("🧑‍⚕️ Give Counselling", use_container_width=True):
            st.session_state.page = "counselling"; st.rerun()
    with c3:
        if st.button("📌 View Schemes", use_container_width=True, type="primary"):
            st.session_state.page = "schemes"; st.rerun()